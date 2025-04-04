import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
from scipy import stats
import os
import openpyxl  # Add this import

def load_data(file_path):
    """Load the CSV data and perform initial checks"""
    try:
        df = pd.read_csv(file_path)
        print(f"Initial data shape: {df.shape}")
        return df
    except FileNotFoundError:
        raise FileNotFoundError(f"Could not find the file at {file_path}")
    except Exception as e:
        raise Exception(f"Error loading data: {str(e)}")

def handle_outliers(df, numeric_cols):
    """Handle outliers using IQR method"""
    for col in numeric_cols:
        Q1 = df[col].quantile(0.25)
        Q3 = df[col].quantile(0.75)
        IQR = Q3 - Q1
        lower_bound = Q1 - 1.5 * IQR
        upper_bound = Q3 + 1.5 * IQR
        df[col] = df[col].clip(lower_bound, upper_bound)
    return df

def clean_data(df):
    """Main data cleaning function"""
    df = df.copy()
    print("\nStarting data cleaning process...")

    # 1. Handle missing values
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].fillna('Unknown')
        elif df[col].dtype in ['int64', 'float64']:
            df[col] = df[col].fillna(df[col].median())
    print("Missing values handled")

    # 2. Clean and standardize text columns
    text_cols = df.select_dtypes(include=['object']).columns
    for col in text_cols:
        df[col] = df[col].str.strip().str.lower()
    print("Text standardized")

    # 3. Handle dates properly
    if 'date_added' in df.columns:
        df['date_added'] = df['date_added'].astype(str).str.strip()
        df['date_added'] = pd.to_datetime(df['date_added'], errors='coerce')

        # Ensure all valid dates are formatted, and replace NaT with empty string
        df['date_added'] = df['date_added'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notnull(x) else '')

        # Extract year and month
        df['year_added'] = pd.to_datetime(df['date_added'], errors='coerce').dt.year
        df['month_added'] = pd.to_datetime(df['date_added'], errors='coerce').dt.month
        print("Dates processed successfully.")

    # 4. Handle duration column
    if 'duration' in df.columns:
        df['duration_value'] = df['duration'].str.extract(r'(\d+)').astype(float)
        df['duration_unit'] = df['duration'].str.extract(r'(\D+)').fillna('min')
        df = df.drop('duration', axis=1)
    print("Duration standardized")

    # 5. Remove duplicates
    df = df.drop_duplicates()
    print("Duplicates removed")

    # 6. Handle outliers in numeric columns
    numeric_cols = df.select_dtypes(include=['int64', 'float64']).columns
    df = handle_outliers(df, numeric_cols)
    print("Outliers handled")

    return df

def save_cleaned_data(df, output_path="cleaned_netflix_data.xlsx"):
    """Save cleaned data in Excel format with proper date formatting for all rows"""
    try:
        if 'date_added' in df.columns:
            df['date_added'] = pd.to_datetime(df['date_added'], errors='coerce')

        with pd.ExcelWriter(output_path, engine='openpyxl', datetime_format='YYYY-MM-DD') as writer:
            df.to_excel(writer, index=False, sheet_name='Netflix Data')
            worksheet = writer.sheets['Netflix Data']

            # Format date column properly
            if 'date_added' in df.columns:
                date_col = df.columns.get_loc('date_added') + 1
                col_letter = openpyxl.utils.get_column_letter(date_col)

                for row in range(2, len(df) + 2):  
                    cell = worksheet.cell(row=row, column=date_col)
                    cell.number_format = 'yyyy-mm-dd'

                worksheet.column_dimensions[col_letter].width = 12

        print(f"\nCleaned data saved to {output_path}")

        # Save sample file with same formatting
        sample_df = df.head(100)
        sample_path = output_path.replace('.xlsx', '_sample.xlsx')
        with pd.ExcelWriter(sample_path, engine='openpyxl', datetime_format='YYYY-MM-DD') as writer:
            sample_df.to_excel(writer, index=False, sheet_name='Netflix Sample')
            worksheet = writer.sheets['Netflix Sample']

            if 'date_added' in sample_df.columns:
                date_col = sample_df.columns.get_loc('date_added') + 1
                col_letter = openpyxl.utils.get_column_letter(date_col)

                for row in range(2, len(sample_df) + 2):
                    cell = worksheet.cell(row=row, column=date_col)
                    cell.number_format = 'yyyy-mm-dd'

                worksheet.column_dimensions[col_letter].width = 12

        print(f"Sample data saved to {sample_path}")

    except Exception as e:
        print(f"Error saving data: {str(e)}")

if __name__ == "__main__":
    try:
        print("Starting data processing...")
        file_path = "D:/Hackathon/netflix_titles.csv"
        df = load_data(file_path)
        print(f"Loaded {len(df)} rows of data")

        print("Cleaning data...")
        cleaned_df = clean_data(df)
        print(f"Cleaned data shape: {cleaned_df.shape}")
        save_cleaned_data(cleaned_df)

    except Exception as e:
        print(f"Error occurred: {str(e)}")
